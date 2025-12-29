"""
Streaming export utilities for FastVue
Supports CSV, Excel, and PDF exports with streaming for large datasets
"""

import csv
import io
from datetime import datetime
from typing import Any, Callable, Dict, List, Optional, Sequence, Type, Union

from fastapi.responses import StreamingResponse
from pydantic import BaseModel


class ExportColumn:
    """Column configuration for exports"""

    def __init__(
        self,
        key: str,
        title: str,
        width: int = 20,
        formatter: Optional[Callable[[Any, Any], str]] = None,
    ):
        self.key = key
        self.title = title
        self.width = width
        self.formatter = formatter

    def get_value(self, obj: Any) -> str:
        """Get formatted value from object"""
        if isinstance(obj, dict):
            value = obj.get(self.key)
        else:
            value = getattr(obj, self.key, None)

        if callable(value):
            value = value()

        if self.formatter:
            return self.formatter(value, obj)

        if value is None:
            return ""

        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d %H:%M:%S")

        return str(value)


def stream_csv_response(
    data: Sequence[Any],
    columns: List[ExportColumn],
    filename: str = "export.csv",
) -> StreamingResponse:
    """
    Stream CSV response for large datasets

    Args:
        data: Sequence of objects or dicts to export
        columns: List of ExportColumn configurations
        filename: Output filename

    Returns:
        StreamingResponse with CSV content
    """
    def generate():
        output = io.StringIO()
        writer = csv.writer(output)

        # Write BOM for UTF-8 Excel compatibility
        yield '\ufeff'

        # Write header
        headers = [col.title for col in columns]
        writer.writerow(headers)
        yield output.getvalue()
        output.seek(0)
        output.truncate(0)

        # Stream data rows
        for obj in data:
            row = [col.get_value(obj) for col in columns]
            writer.writerow(row)
            yield output.getvalue()
            output.seek(0)
            output.truncate(0)

    return StreamingResponse(
        generate(),
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Cache-Control": "no-cache",
        },
    )


def export_to_csv(
    data: Sequence[Any],
    columns: List[ExportColumn],
) -> str:
    """
    Export data to CSV string

    Args:
        data: Sequence of objects or dicts to export
        columns: List of ExportColumn configurations

    Returns:
        CSV string content
    """
    output = io.StringIO()
    writer = csv.writer(output)

    # Write BOM for UTF-8 Excel compatibility
    output.write('\ufeff')

    # Write header
    headers = [col.title for col in columns]
    writer.writerow(headers)

    # Write data rows
    for obj in data:
        row = [col.get_value(obj) for col in columns]
        writer.writerow(row)

    return output.getvalue()


def export_to_excel(
    data: Sequence[Any],
    columns: List[ExportColumn],
    sheet_name: str = "Export",
) -> bytes:
    """
    Export data to Excel bytes

    Args:
        data: Sequence of objects or dicts to export
        columns: List of ExportColumn configurations
        sheet_name: Excel sheet name

    Returns:
        Excel file as bytes
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    # Write headers
    for col_idx, column in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=column.title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = column.width

    # Write data
    for row_idx, obj in enumerate(data, 2):
        for col_idx, column in enumerate(columns, 1):
            ws.cell(row=row_idx, column=col_idx, value=column.get_value(obj))

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def stream_excel_response(
    data: Sequence[Any],
    columns: List[ExportColumn],
    filename: str = "export.xlsx",
    sheet_name: str = "Export",
) -> StreamingResponse:
    """
    Stream Excel response

    Args:
        data: Sequence of objects or dicts to export
        columns: List of ExportColumn configurations
        filename: Output filename
        sheet_name: Excel sheet name

    Returns:
        StreamingResponse with Excel content
    """
    excel_bytes = export_to_excel(data, columns, sheet_name)

    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Cache-Control": "no-cache",
        },
    )


# Common formatters
def format_date(value: Any, obj: Any = None) -> str:
    """Format date value"""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value)


def format_datetime(value: Any, obj: Any = None) -> str:
    """Format datetime value"""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return str(value)


def format_boolean(value: Any, obj: Any = None) -> str:
    """Format boolean value"""
    if value is None:
        return ""
    return "Yes" if value else "No"


def format_currency(value: Any, obj: Any = None, symbol: str = "$") -> str:
    """Format currency value"""
    if value is None:
        return ""
    try:
        return f"{symbol}{float(value):,.2f}"
    except (ValueError, TypeError):
        return str(value)


def format_percentage(value: Any, obj: Any = None) -> str:
    """Format percentage value"""
    if value is None:
        return ""
    try:
        return f"{float(value):.1f}%"
    except (ValueError, TypeError):
        return str(value)


# Predefined export columns for common models
def create_user_export_columns() -> List[ExportColumn]:
    """Create export columns for User model"""
    return [
        ExportColumn("id", "ID", 10),
        ExportColumn("username", "Username", 20),
        ExportColumn("email", "Email", 30),
        ExportColumn("full_name", "Full Name", 25),
        ExportColumn("is_active", "Active", 10, format_boolean),
        ExportColumn("is_superuser", "Admin", 10, format_boolean),
        ExportColumn("created_at", "Created", 20, format_datetime),
        ExportColumn("last_login_at", "Last Login", 20, format_datetime),
    ]


def create_company_export_columns() -> List[ExportColumn]:
    """Create export columns for Company model"""
    return [
        ExportColumn("id", "ID", 10),
        ExportColumn("code", "Code", 15),
        ExportColumn("name", "Company Name", 30),
        ExportColumn("city", "City", 20),
        ExportColumn("country", "Country", 20),
        ExportColumn("is_active", "Active", 10, format_boolean),
        ExportColumn("is_headquarters", "HQ", 10, format_boolean),
        ExportColumn("created_at", "Created", 20, format_datetime),
    ]


def create_audit_export_columns() -> List[ExportColumn]:
    """Create export columns for AuditLog model"""
    return [
        ExportColumn("id", "ID", 10),
        ExportColumn("user_name", "User", 20),
        ExportColumn("action", "Action", 15),
        ExportColumn("entity_type", "Entity", 20),
        ExportColumn("entity_id", "Entity ID", 10),
        ExportColumn("ip_address", "IP Address", 15),
        ExportColumn("created_at", "Timestamp", 20, format_datetime),
    ]


# Bulk operations with chunking
def bulk_process_with_chunks(
    items: Sequence[Any],
    processor: Callable[[Sequence[Any]], None],
    chunk_size: int = 1000,
) -> int:
    """
    Process items in chunks to avoid memory issues

    Args:
        items: Sequence of items to process
        processor: Function to process each chunk
        chunk_size: Size of each chunk

    Returns:
        Total number of items processed
    """
    total = 0
    for i in range(0, len(items), chunk_size):
        chunk = items[i:i + chunk_size]
        processor(chunk)
        total += len(chunk)
    return total
