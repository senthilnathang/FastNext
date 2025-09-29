from typing import Dict, List, Any, Optional, Union, AsyncGenerator, BinaryIO, IO
from fastapi import BackgroundTasks, HTTPException, status, UploadFile
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import text, select, func, MetaData, Table, Column, inspect
from sqlalchemy.orm import selectinload
import asyncio
import csv
import json
import xml.etree.ElementTree as ET
import pandas as pd
import io
import logging
from datetime import datetime, date
from decimal import Decimal
from pathlib import Path
import tempfile
import aiofiles
from enum import Enum
import uuid
import re
from typing import AsyncGenerator
import yaml
import openpyxl
from openpyxl import load_workbook

from app.core.config import settings
from app.core.logging import get_logger
from app.models.base import Base
from app.models.data_import_export import ImportJob, ImportStatus, DataFormat
from app.schemas.data_import_export import (
    ImportOptionsSchema, 
    FieldMappingSchema,
    ValidationErrorSchema,
    ValidationResultSchema
)

logger = get_logger(__name__)


class ImportError(Exception):
    """Custom exception for import errors"""
    pass


class ValidationError(ImportError):
    """Custom exception for validation errors"""
    pass


class DataImporter:
    """Comprehensive data import utility with multiple format support"""
    
    def __init__(self, session: AsyncSession):
        self.session = session
        self.import_jobs: Dict[str, Dict[str, Any]] = {}
        self.max_file_size = 100 * 1024 * 1024  # 100MB limit
        self.max_rows_per_import = 1000000  # 1M rows limit
        self.chunk_size = 1000  # Process in chunks
        self.supported_formats = [DataFormat.CSV, DataFormat.JSON, DataFormat.EXCEL, DataFormat.XML, DataFormat.YAML]
    
    async def validate_file(self, file: UploadFile, max_size: Optional[int] = None) -> bool:
        """Validate uploaded file"""
        
        # Check file size
        max_allowed_size = max_size or self.max_file_size
        file.file.seek(0, 2)  # Seek to end
        file_size = file.file.tell()
        file.file.seek(0)  # Reset to beginning
        
        if file_size > max_allowed_size:
            raise ImportError(f"File too large. Maximum size: {max_allowed_size / (1024*1024):.1f}MB")
        
        # Check file format by extension
        if not file.filename:
            raise ImportError("Filename is required")
            
        file_ext = Path(file.filename).suffix.lower()
        format_extensions = {
            '.csv': DataFormat.CSV,
            '.json': DataFormat.JSON,
            '.xlsx': DataFormat.EXCEL,
            '.xls': DataFormat.EXCEL,
            '.xml': DataFormat.XML,
            '.yaml': DataFormat.YAML,
            '.yml': DataFormat.YAML
        }
        
        if file_ext not in format_extensions:
            raise ImportError(f"Unsupported file format: {file_ext}")
        
        return True
    
    async def parse_file(
        self,
        file: UploadFile,
        import_options: ImportOptionsSchema
    ) -> Dict[str, Any]:
        """Parse uploaded file and return structured data"""
        
        try:
            await self.validate_file(file)
            
            # Determine format
            file_format = import_options.format
            if not file_format:
                file_format = self._detect_format_from_filename(file.filename)
            
            # Read file content
            content = await file.read()
            
            # Parse based on format
            if file_format == DataFormat.CSV:
                return await self._parse_csv(content, import_options)
            elif file_format == DataFormat.JSON:
                return await self._parse_json(content, import_options)
            elif file_format == DataFormat.EXCEL:
                return await self._parse_excel(content, import_options)
            elif file_format == DataFormat.XML:
                return await self._parse_xml(content, import_options)
            elif file_format == DataFormat.YAML:
                return await self._parse_yaml(content, import_options)
            else:
                raise ImportError(f"Unsupported format: {file_format}")
                
        except Exception as e:
            logger.error(f"File parsing failed: {e}")
            raise ImportError(f"Failed to parse file: {str(e)}")
    
    async def _parse_csv(self, content: bytes, options: ImportOptionsSchema) -> Dict[str, Any]:
        """Parse CSV file content"""
        
        try:
            # Decode content
            text_content = content.decode(options.encoding or 'utf-8')
            
            # Create CSV reader
            csv_file = io.StringIO(text_content)
            delimiter = options.delimiter or ','
            reader = csv.reader(csv_file, delimiter=delimiter)
            
            rows = list(reader)
            
            if not rows:
                raise ImportError("CSV file is empty")
            
            # Extract headers
            if options.has_headers:
                headers = rows[0]
                data_rows = rows[1:]
            else:
                headers = [f"column_{i+1}" for i in range(len(rows[0]))]
                data_rows = rows
            
            # Skip rows if specified
            if options.skip_first_rows > 0:
                data_rows = data_rows[options.skip_first_rows:]
            
            # Convert to dictionaries
            structured_data = []
            for i, row in enumerate(data_rows):
                if options.skip_empty_rows and not any(cell.strip() for cell in row):
                    continue
                
                row_dict = {}
                for j, value in enumerate(row):
                    if j < len(headers):
                        row_dict[headers[j]] = value.strip() if value else None
                
                structured_data.append(row_dict)
                
                # Respect max rows limit
                if options.max_rows and len(structured_data) >= options.max_rows:
                    break
            
            return {
                "headers": headers,
                "rows": structured_data,
                "total_rows": len(structured_data),
                "format": DataFormat.CSV
            }
            
        except UnicodeDecodeError:
            raise ImportError(f"Could not decode file with encoding: {options.encoding}")
        except Exception as e:
            raise ImportError(f"CSV parsing error: {str(e)}")
    
    async def _parse_json(self, content: bytes, options: ImportOptionsSchema) -> Dict[str, Any]:
        """Parse JSON file content"""
        
        try:
            # Decode and parse JSON
            text_content = content.decode(options.encoding or 'utf-8')
            data = json.loads(text_content)
            
            # Handle different JSON structures
            if isinstance(data, list):
                structured_data = data
            elif isinstance(data, dict):
                # Look for common array keys
                if 'data' in data:
                    structured_data = data['data']
                elif 'rows' in data:
                    structured_data = data['rows']
                elif 'records' in data:
                    structured_data = data['records']
                else:
                    # Treat as single record
                    structured_data = [data]
            else:
                raise ImportError("JSON must contain an object or array")
            
            if not structured_data:
                raise ImportError("JSON file contains no data")
            
            # Extract headers from first record
            if structured_data and isinstance(structured_data[0], dict):
                headers = list(structured_data[0].keys())
            else:
                raise ImportError("JSON data must be an array of objects")
            
            # Apply row limits
            if options.max_rows:
                structured_data = structured_data[:options.max_rows]
            
            return {
                "headers": headers,
                "rows": structured_data,
                "total_rows": len(structured_data),
                "format": DataFormat.JSON
            }
            
        except json.JSONDecodeError as e:
            raise ImportError(f"Invalid JSON: {str(e)}")
        except Exception as e:
            raise ImportError(f"JSON parsing error: {str(e)}")
    
    async def _parse_excel(self, content: bytes, options: ImportOptionsSchema) -> Dict[str, Any]:
        """Parse Excel file content"""
        
        try:
            # Load workbook from bytes
            with io.BytesIO(content) as excel_file:
                workbook = openpyxl.load_workbook(excel_file, data_only=True)
            
            # Use first worksheet
            worksheet = workbook.active
            
            # Get all rows
            rows = list(worksheet.iter_rows(values_only=True))
            
            if not rows:
                raise ImportError("Excel file is empty")
            
            # Extract headers
            if options.has_headers:
                headers = [str(cell) if cell is not None else f"column_{i+1}" for i, cell in enumerate(rows[0])]
                data_rows = rows[1:]
            else:
                headers = [f"column_{i+1}" for i in range(len(rows[0]))]
                data_rows = rows
            
            # Skip rows if specified
            if options.skip_first_rows > 0:
                data_rows = data_rows[options.skip_first_rows:]
            
            # Convert to dictionaries
            structured_data = []
            for row in data_rows:
                if options.skip_empty_rows and not any(cell for cell in row):
                    continue
                
                row_dict = {}
                for j, value in enumerate(row):
                    if j < len(headers):
                        # Convert Excel values to strings
                        if value is not None:
                            if isinstance(value, datetime):
                                row_dict[headers[j]] = value.isoformat()
                            else:
                                row_dict[headers[j]] = str(value)
                        else:
                            row_dict[headers[j]] = None
                
                structured_data.append(row_dict)
                
                # Respect max rows limit
                if options.max_rows and len(structured_data) >= options.max_rows:
                    break
            
            return {
                "headers": headers,
                "rows": structured_data,
                "total_rows": len(structured_data),
                "format": DataFormat.EXCEL
            }
            
        except Exception as e:
            raise ImportError(f"Excel parsing error: {str(e)}")
    
    async def _parse_xml(self, content: bytes, options: ImportOptionsSchema) -> Dict[str, Any]:
        """Parse XML file content"""
        
        try:
            # Parse XML
            root = ET.fromstring(content)
            
            # Find repeating elements (assume these are data rows)
            # Look for common patterns
            data_elements = []
            
            # Try common XML structures
            for child in root:
                if len(list(child)) > 0:  # Element has children
                    data_elements.append(child)
            
            if not data_elements:
                # If no child elements, try using root children as records
                data_elements = list(root)
            
            if not data_elements:
                raise ImportError("No data elements found in XML")
            
            # Extract headers from first element
            first_element = data_elements[0]
            headers = []
            
            # Get all child element tags as headers
            for child in first_element:
                headers.append(child.tag)
            
            # If no child elements, use attributes as headers
            if not headers:
                headers = list(first_element.attrib.keys())
            
            if not headers:
                raise ImportError("Could not determine data structure from XML")
            
            # Convert to structured data
            structured_data = []
            for element in data_elements:
                row_dict = {}
                
                # Extract data from child elements
                for child in element:
                    row_dict[child.tag] = child.text
                
                # Extract data from attributes if no child elements
                if not list(element):
                    for key, value in element.attrib.items():
                        row_dict[key] = value
                
                structured_data.append(row_dict)
                
                # Respect max rows limit
                if options.max_rows and len(structured_data) >= options.max_rows:
                    break
            
            return {
                "headers": headers,
                "rows": structured_data,
                "total_rows": len(structured_data),
                "format": DataFormat.XML
            }
            
        except ET.ParseError as e:
            raise ImportError(f"Invalid XML: {str(e)}")
        except Exception as e:
            raise ImportError(f"XML parsing error: {str(e)}")
    
    async def _parse_yaml(self, content: bytes, options: ImportOptionsSchema) -> Dict[str, Any]:
        """Parse YAML file content"""
        
        try:
            # Decode and parse YAML
            text_content = content.decode(options.encoding or 'utf-8')
            data = yaml.safe_load(text_content)
            
            # Handle different YAML structures
            if isinstance(data, list):
                structured_data = data
            elif isinstance(data, dict):
                # Look for common array keys
                if 'data' in data:
                    structured_data = data['data']
                elif 'rows' in data:
                    structured_data = data['rows']
                elif 'records' in data:
                    structured_data = data['records']
                else:
                    # Treat as single record
                    structured_data = [data]
            else:
                raise ImportError("YAML must contain an object or array")
            
            if not structured_data:
                raise ImportError("YAML file contains no data")
            
            # Extract headers from first record
            if structured_data and isinstance(structured_data[0], dict):
                headers = list(structured_data[0].keys())
            else:
                raise ImportError("YAML data must be an array of objects")
            
            # Apply row limits
            if options.max_rows:
                structured_data = structured_data[:options.max_rows]
            
            return {
                "headers": headers,
                "rows": structured_data,
                "total_rows": len(structured_data),
                "format": DataFormat.YAML
            }
            
        except yaml.YAMLError as e:
            raise ImportError(f"Invalid YAML: {str(e)}")
        except Exception as e:
            raise ImportError(f"YAML parsing error: {str(e)}")
    
    def _detect_format_from_filename(self, filename: str) -> DataFormat:
        """Detect file format from filename extension"""
        
        file_ext = Path(filename).suffix.lower()
        format_map = {
            '.csv': DataFormat.CSV,
            '.json': DataFormat.JSON,
            '.xlsx': DataFormat.EXCEL,
            '.xls': DataFormat.EXCEL,
            '.xml': DataFormat.XML,
            '.yaml': DataFormat.YAML,
            '.yml': DataFormat.YAML
        }
        
        return format_map.get(file_ext, DataFormat.CSV)
    
    async def validate_data(
        self,
        data: List[Dict[str, Any]],
        field_mappings: List[FieldMappingSchema],
        table_name: str
    ) -> ValidationResultSchema:
        """Validate parsed data against target table schema"""
        
        try:
            # Get table schema
            table_schema = await self._get_table_schema(table_name)
            
            errors = []
            warnings = []
            valid_rows = 0
            
            # Create mapping lookup
            mapping_lookup = {mapping.source_column: mapping for mapping in field_mappings}
            
            for row_idx, row in enumerate(data, 1):
                row_errors = []
                
                # Validate each mapped field
                for source_col, mapping in mapping_lookup.items():
                    if source_col in row:
                        value = row[source_col]
                        
                        # Get target column info
                        target_col_info = table_schema.get(mapping.target_column)
                        if not target_col_info:
                            row_errors.append(ValidationErrorSchema(
                                row=row_idx,
                                column=source_col,
                                field=mapping.target_column,
                                message=f"Target column '{mapping.target_column}' does not exist",
                                severity="error"
                            ))
                            continue
                        
                        # Validate value
                        validation_error = self._validate_field_value(
                            value, target_col_info, row_idx, source_col, mapping.target_column
                        )
                        if validation_error:
                            row_errors.append(validation_error)
                
                # Check for required fields
                for target_col, col_info in table_schema.items():
                    if col_info.get('nullable') is False and col_info.get('default') is None:
                        # Find if this required field is mapped
                        mapped = False
                        for mapping in field_mappings:
                            if mapping.target_column == target_col and mapping.source_column in row:
                                if row[mapping.source_column] is not None and str(row[mapping.source_column]).strip():
                                    mapped = True
                                    break
                        
                        if not mapped:
                            row_errors.append(ValidationErrorSchema(
                                row=row_idx,
                                field=target_col,
                                message=f"Required field '{target_col}' is missing or empty",
                                severity="error"
                            ))
                
                if row_errors:
                    errors.extend(row_errors)
                else:
                    valid_rows += 1
            
            return ValidationResultSchema(
                is_valid=len(errors) == 0,
                total_rows=len(data),
                valid_rows=valid_rows,
                error_rows=len(data) - valid_rows,
                errors=errors,
                warnings=warnings
            )
            
        except Exception as e:
            logger.error(f"Data validation failed: {e}")
            raise ValidationError(f"Validation failed: {str(e)}")
    
    async def _get_table_schema(self, table_name: str) -> Dict[str, Any]:
        """Get table schema information"""
        
        try:
            # Use SQLAlchemy inspector to get table info
            inspector = inspect(self.session.bind)
            columns = inspector.get_columns(table_name)
            
            schema = {}
            for col in columns:
                schema[col['name']] = {
                    'type': str(col['type']),
                    'nullable': col['nullable'],
                    'default': col['default'],
                    'primary_key': col.get('primary_key', False),
                    'autoincrement': col.get('autoincrement', False)
                }
            
            return schema
            
        except Exception as e:
            raise ImportError(f"Could not get schema for table '{table_name}': {str(e)}")
    
    def _validate_field_value(
        self,
        value: Any,
        column_info: Dict[str, Any],
        row_idx: int,
        source_column: str,
        target_column: str
    ) -> Optional[ValidationErrorSchema]:
        """Validate a single field value"""
        
        # Check for null values
        if value is None or (isinstance(value, str) and not value.strip()):
            if not column_info['nullable']:
                return ValidationErrorSchema(
                    row=row_idx,
                    column=source_column,
                    field=target_column,
                    message=f"Field cannot be null",
                    severity="error"
                )
            return None
        
        # Type validation based on column type
        column_type = column_info['type'].lower()
        
        try:
            if 'int' in column_type:
                int(value)
            elif 'float' in column_type or 'decimal' in column_type or 'numeric' in column_type:
                float(value)
            elif 'bool' in column_type:
                if str(value).lower() not in ['true', 'false', '1', '0', 'yes', 'no']:
                    return ValidationErrorSchema(
                        row=row_idx,
                        column=source_column,
                        field=target_column,
                        message=f"Invalid boolean value: {value}",
                        severity="error"
                    )
            elif 'date' in column_type:
                # Try to parse date
                try:
                    datetime.fromisoformat(str(value).replace('Z', '+00:00'))
                except:
                    return ValidationErrorSchema(
                        row=row_idx,
                        column=source_column,
                        field=target_column,
                        message=f"Invalid date format: {value}",
                        severity="error"
                    )
        
        except (ValueError, TypeError):
            return ValidationErrorSchema(
                row=row_idx,
                column=source_column,
                field=target_column,
                message=f"Invalid {column_type} value: {value}",
                severity="error"
            )
        
        return None
    
    async def import_data(
        self,
        data: List[Dict[str, Any]],
        field_mappings: List[FieldMappingSchema],
        table_name: str,
        options: ImportOptionsSchema,
        job_id: Optional[str] = None
    ) -> Dict[str, Any]:
        """Import validated data into database"""
        
        try:
            imported_rows = 0
            error_rows = 0
            errors = []
            
            # Create mapping lookup
            mapping_lookup = {mapping.source_column: mapping for mapping in field_mappings}
            
            # Process in chunks
            for chunk_start in range(0, len(data), self.chunk_size):
                chunk_end = min(chunk_start + self.chunk_size, len(data))
                chunk_data = data[chunk_start:chunk_end]
                
                # Transform chunk data according to mappings
                transformed_rows = []
                for row_idx, row in enumerate(chunk_data):
                    transformed_row = {}
                    
                    for source_col, value in row.items():
                        if source_col in mapping_lookup:
                            mapping = mapping_lookup[source_col]
                            target_col = mapping.target_column
                            
                            # Apply transformation if specified
                            if mapping.transform:
                                value = self._apply_transform(value, mapping.transform)
                            
                            # Skip empty values if configured
                            if mapping.skip_empty and (value is None or str(value).strip() == ""):
                                continue
                            
                            transformed_row[target_col] = value
                    
                    if transformed_row:  # Only add non-empty rows
                        transformed_rows.append(transformed_row)
                
                # Insert chunk into database
                if transformed_rows:
                    try:
                        await self._insert_chunk(table_name, transformed_rows, options)
                        imported_rows += len(transformed_rows)
                    except Exception as e:
                        error_rows += len(transformed_rows)
                        errors.append(f"Chunk {chunk_start}-{chunk_end}: {str(e)}")
                        
                        if options.validate_only:
                            # Don't continue if this is just validation
                            break
            
            return {
                "imported_rows": imported_rows,
                "error_rows": error_rows,
                "total_rows": len(data),
                "errors": errors,
                "success_rate": imported_rows / len(data) if data else 0
            }
            
        except Exception as e:
            logger.error(f"Data import failed: {e}")
            raise ImportError(f"Import failed: {str(e)}")
    
    async def _insert_chunk(
        self,
        table_name: str,
        rows: List[Dict[str, Any]],
        options: ImportOptionsSchema
    ):
        """Insert a chunk of data into the database"""
        
        if not rows:
            return
        
        try:
            # Handle duplicate strategy
            if options.on_duplicate == "skip":
                # Use INSERT IGNORE equivalent or ON CONFLICT DO NOTHING
                # This is PostgreSQL specific - adjust for other databases
                placeholders = ", ".join([f"({', '.join([':' + k + str(i) for k in row.keys()])})" 
                                        for i, row in enumerate(rows)])
                columns = ", ".join(rows[0].keys())
                
                query = f"""
                INSERT INTO {table_name} ({columns}) 
                VALUES {placeholders}
                ON CONFLICT DO NOTHING
                """
                
                # Flatten parameters
                params = {}
                for i, row in enumerate(rows):
                    for k, v in row.items():
                        params[f"{k}{i}"] = v
                
                await self.session.execute(text(query), params)
                
            elif options.on_duplicate == "update":
                # Handle updates - more complex, simplified here
                for row in rows:
                    await self._upsert_row(table_name, row)
                    
            else:  # "error" - default insert
                # Simple insert all
                for row in rows:
                    columns = ", ".join(row.keys())
                    placeholders = ", ".join([f":{k}" for k in row.keys()])
                    query = f"INSERT INTO {table_name} ({columns}) VALUES ({placeholders})"
                    
                    await self.session.execute(text(query), row)
            
            await self.session.commit()
            
        except Exception as e:
            await self.session.rollback()
            raise e
    
    async def _upsert_row(self, table_name: str, row: Dict[str, Any]):
        """Upsert a single row (PostgreSQL specific)"""
        
        columns = ", ".join(row.keys())
        placeholders = ", ".join([f":{k}" for k in row.keys()])
        
        # Simple upsert - assumes 'id' as primary key
        # In production, you'd want to make this more flexible
        update_clause = ", ".join([f"{k} = EXCLUDED.{k}" for k in row.keys() if k != 'id'])
        
        query = f"""
        INSERT INTO {table_name} ({columns}) 
        VALUES ({placeholders})
        ON CONFLICT (id) DO UPDATE SET {update_clause}
        """
        
        await self.session.execute(text(query), row)
    
    def _apply_transform(self, value: Any, transform: str) -> Any:
        """Apply data transformation"""
        
        if not value or not transform:
            return value
        
        try:
            if transform == "upper":
                return str(value).upper()
            elif transform == "lower":
                return str(value).lower()
            elif transform == "trim":
                return str(value).strip()
            elif transform == "date":
                # Try to parse and format date
                try:
                    dt = datetime.fromisoformat(str(value).replace('Z', '+00:00'))
                    return dt.strftime('%Y-%m-%d')
                except:
                    return value
            elif transform == "number":
                try:
                    return float(value)
                except:
                    return 0
            elif transform == "boolean":
                return str(value).lower() in ['true', '1', 'yes', 'on']
            else:
                return value
        except:
            return value


# Global importer instance
data_importer = None

def get_data_importer(session: AsyncSession) -> DataImporter:
    """Get or create data importer instance"""
    global data_importer
    if data_importer is None:
        data_importer = DataImporter(session)
    return data_importer